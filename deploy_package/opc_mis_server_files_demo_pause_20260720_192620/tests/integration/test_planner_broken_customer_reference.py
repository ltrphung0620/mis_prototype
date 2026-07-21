"""Broken contract-to-customer reference integration behavior."""

import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from opc_mis.domain.enums import (
    ReadinessStatus,
    SourceType,
    WorkflowStatus,
)
from opc_mis.domain.evidence import DataPatch
from opc_mis.domain.team_pack import SheetRegistry
from tests.conftest import execute_planner, make_request


def test_broken_contract_customer_reference_is_blocking(
    team_pack_path: Path, tmp_path: Path
) -> None:
    copied = tmp_path / "broken_customer.xlsx"
    shutil.copyfile(team_pack_path, copied)
    workbook = load_workbook(copied)
    contracts = workbook[SheetRegistry.CONTRACTS.sheet_name]
    headers = [cell.value for cell in contracts[contracts.min_row]]
    contract_column = headers.index("contract_id") + 1
    customer_column = headers.index("customer_id") + 1
    selected_row = next(
        row
        for row in range(2, contracts.max_row + 1)
        if contracts.cell(row=row, column=contract_column).value
    )
    contract_id = str(contracts.cell(row=selected_row, column=contract_column).value)
    contracts.cell(row=selected_row, column=customer_column, value="CUS-NOT-PRESENT")
    workbook.save(copied)

    result = execute_planner(make_request(copied, contract_id))

    assert result.status is WorkflowStatus.WAITING_FOR_INPUT
    assert result.planner_result is not None
    assert result.planner_result.data_readiness.status is ReadinessStatus.BLOCKED
    assert any(
        item.requirement_code == "BROKEN_CONTRACT_CUSTOMER_REFERENCE"
        for item in result.planner_result.missing_data_requests
    )


def test_duplicate_related_invoice_pauses_the_case(
    team_pack_path: Path, tmp_path: Path, first_contract_id: str
) -> None:
    orders = pd.read_excel(team_pack_path, sheet_name="06_ORDERS", dtype=object)
    invoices = pd.read_excel(team_pack_path, sheet_name="07_INVOICES", dtype=object)
    related_order_ids = set(
        orders.loc[orders["contract_id"] == first_contract_id, "order_id"].astype(str)
    )
    invoice_id = str(
        invoices.loc[invoices["order_id"].isin(related_order_ids), "invoice_id"].iloc[0]
    )
    copied = tmp_path / "duplicate_related_invoice.xlsx"
    shutil.copyfile(team_pack_path, copied)
    workbook = load_workbook(copied)
    invoice_sheet = workbook[SheetRegistry.INVOICES.sheet_name]
    headers = [cell.value for cell in invoice_sheet[invoice_sheet.min_row]]
    identifier_column = headers.index("invoice_id") + 1
    selected_row = next(
        row
        for row in range(invoice_sheet.min_row + 1, invoice_sheet.max_row + 1)
        if str(invoice_sheet.cell(row=row, column=identifier_column).value) == invoice_id
    )
    invoice_sheet.append([cell.value for cell in invoice_sheet[selected_row]])
    workbook.save(copied)

    result = execute_planner(make_request(copied, first_contract_id))

    assert result.status is WorkflowStatus.WAITING_FOR_INPUT
    assert result.planner_result is not None
    assert any(
        item.requirement_code == "DUPLICATE_PRIMARY_KEY" and item.target_record == invoice_id
        for item in result.planner_result.missing_data_requests
    )
    assert result.planner_result.run_plan.parallel_initial_tasks == ()


def test_patch_created_foreign_key_break_is_revalidated(
    team_pack_path: Path, first_contract_id: str
) -> None:
    orders = pd.read_excel(team_pack_path, sheet_name="06_ORDERS", dtype=object)
    order_id = str(orders.loc[orders["contract_id"] == first_contract_id, "order_id"].iloc[0])
    patch = DataPatch(
        patch_id="PATCH-BROKEN-SERVICE-FK",
        source=SourceType.USER_INPUT,
        canonical_entity_type="ORDER",
        target_record=order_id,
        field="service_id",
        value="SVC-NOT-PRESENT",
        evidence_note="Verify overlay relationship revalidation",
    )

    result = execute_planner(make_request(team_pack_path, first_contract_id, data_patches=(patch,)))

    assert result.status is WorkflowStatus.WAITING_FOR_INPUT
    assert result.planner_result is not None
    assert any(
        item.requirement_code == "BROKEN_FOREIGN_KEY"
        and item.target_record == order_id
        and item.field == "service_id"
        for item in result.planner_result.missing_data_requests
    )
    assert result.planner_result.run_plan.parallel_initial_tasks == ()
