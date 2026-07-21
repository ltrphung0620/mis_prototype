"""Unit tests for actual TeamPack workbook ingestion."""

import json
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from opc_mis.domain.team_pack import SheetRegistry
from opc_mis.infrastructure.excel.normalizers import json_safe
from opc_mis.infrastructure.excel.workbook_loader import WorkbookLoader


def test_loader_uses_actual_sheet_names(team_pack_path: Path) -> None:
    dataset = WorkbookLoader().load("DATASET", team_pack_path)

    assert SheetRegistry.CONTRACTS.sheet_name in dataset.sheets
    assert SheetRegistry.ORDERS.sheet_name in dataset.sheets
    assert dataset.headers[SheetRegistry.CONTRACTS.sheet_name][0] == "contract_id"
    assert not dataset.missing_sheets


def test_nan_and_nat_are_json_safe() -> None:
    assert json_safe(float("nan")) is None
    assert json.dumps({"value": json_safe(float("nan"))}) == '{"value": null}'


def test_nested_pandas_values_are_json_safe() -> None:
    payload = json_safe({"values": [float("nan"), pd.NA, pd.NaT]})

    encoded = json.dumps(payload, allow_nan=False)

    assert encoded == '{"values": [null, null, null]}'


def test_duplicate_primary_keys_are_detected(team_pack_path: Path, tmp_path: Path) -> None:
    copied = tmp_path / "duplicate.xlsx"
    shutil.copyfile(team_pack_path, copied)
    workbook = load_workbook(copied)
    sheet = workbook[SheetRegistry.CUSTOMERS.sheet_name]
    headers = [cell.value for cell in sheet[sheet.min_row]]
    identifier_column = headers.index("customer_id") + 1
    selected_row = next(
        row
        for row in range(sheet.min_row + 1, sheet.max_row + 1)
        if sheet.cell(row=row, column=identifier_column).value
    )
    duplicate_values = [cell.value for cell in sheet[selected_row]]
    sheet.append(duplicate_values)
    duplicate_id = str(duplicate_values[0])
    workbook.save(copied)

    dataset = WorkbookLoader().load("DATASET", copied)

    assert duplicate_id in dataset.duplicate_ids[SheetRegistry.CUSTOMERS.sheet_name]


def test_explicit_foreign_key_break_is_reported(team_pack_path: Path, tmp_path: Path) -> None:
    copied = tmp_path / "broken_fk.xlsx"
    shutil.copyfile(team_pack_path, copied)
    workbook = load_workbook(copied)
    sheet = workbook[SheetRegistry.CONTRACTS.sheet_name]
    headers = [cell.value for cell in sheet[sheet.min_row]]
    identifier_column = headers.index("contract_id") + 1
    customer_column = headers.index("customer_id") + 1
    selected_row = next(
        row
        for row in range(sheet.min_row + 1, sheet.max_row + 1)
        if sheet.cell(row=row, column=identifier_column).value
    )
    sheet.cell(row=selected_row, column=customer_column, value="CUS-NOT-PRESENT")
    workbook.save(copied)

    dataset = WorkbookLoader().load("DATASET", copied)

    assert any(
        issue.code == "BROKEN_FOREIGN_KEY" and issue.sheet == SheetRegistry.CONTRACTS.sheet_name
        for issue in dataset.validation_issues
    )
